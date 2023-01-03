import pandas as pd
import openpyxl as op
from datetime import datetime
from pprint import pprint


def initializer():
    df_items = pd.read_excel(io='data_files/test.xlsx', sheet_name='Inventory')
    df_sales = pd.read_excel(io='data_files/test.xlsx', sheet_name='Sales')

    if df_items.columns.size == 0:
        item = [{"Item Desc": "", "Category": "", "Made in": "", "Size/ml-g-oz": "", "Unit": "", "Quantity": "",
                 "Bar Code": "", "Unit Price": "", "45% Profit": "", "Unit Price after profit": "", "VAT": "",
                 "Total": "", "On hand qty": "", "Sold qty": "", "Total sales": ""}]

        for ite in item[0].keys():
            item[0][ite] = ite

        df_item = pd.DataFrame(item)
        rows = df_item.values.tolist()
        book = op.load_workbook(filename='data_files/test.xlsx')
        sheet = book['Inventory']

        for row in rows:
            sheet.append(row)

        book.save(filename='data_files/test.xlsx')

        print("Inventory sheet column labels created")
    else:
        print("Inventory sheet already has column labels")

    if df_sales.columns.size == 0:
        sale = [{"Item Desc": "", "Price": "", "Quantity": "", "Total": "", "Customer": "", "Time": ""}]

        for ite in sale[0].keys():
            sale[0][ite] = ite

        df_sale = pd.DataFrame(sale)
        rows = df_sale.values.tolist()
        book = op.load_workbook(filename='data_files/test.xlsx')
        sheet = book['Sales']

        for row in rows:
            sheet.append(row)

        book.save(filename='data_files/test.xlsx')

        print("Sales sheet column labels created")
    else:
        print("Sales sheet already has column labels")


initializer()


def fetchItem(name: str):
    df = pd.read_excel('data_files/test.xlsx', sheet_name='Inventory')
    # print(df)
    if name in (df.loc[:, 'Item Desc']).to_list():
        idx = (df.loc[:, 'Item Desc']).to_list().index(name)
        item = {"id": idx, "name": df.loc[:, 'Item Desc'].to_list()[idx],
                "category": df.loc[:, 'Category'].to_list()[idx], "made in": df.loc[:, 'Made in'].to_list()[idx],
                "size": df.loc[:, 'Size/ml-g-oz'].to_list()[idx], "price": df.loc[:, 'Unit Price'].to_list()[idx],
                "profit": df.loc[:, 'Profit'].to_list()[idx], "available": df.loc[:, 'On hand qty'].to_list()[idx]}
        print(item, "\n")
        return item
    else:
        print(df)
        return -1
# fetchItem(10)


def fetchAll():
    df = pd.read_excel('data_files/test.xlsx', sheet_name='Inventory')
    dic = df.loc[:].to_dict()
    la = {}
    li = {}
    # k = open('kk.json', 'w')
    for i in range(0, len(df)):
        li[i] = {}
        for key in dic.keys():
            if len(li[i]) == 0:
                li[i] = {key: ""}
            li[i][key] = dic[key][i]
    return li


# pprint(fetchAll(), indent=4)

def fetchAllSold():
    df = pd.read_excel('data_files/test.xlsx', sheet_name='Sales')
    dic = df.loc[:].to_dict()
    la = {}
    li = {}
    # k = open('kk.json', 'w')
    for i in range(0, len(df)):
        li[i] = {}
        for key in dic.keys():
            if len(li[i]) == 0:
                li[i] = {key: ""}
            li[i][key] = dic[key][i]
    return li


# fetchAllSold()

def calculateSales(price, quantity: int):
    total = int(quantity) * price
    print(total)
    return total


# calculateSales(10, 20)

def modify_after_sale(item_name: str, quantity: int):
    item = fetchItem(item_name)
    df_items = pd.read_excel(io='data_files/test.xlsx', sheet_name='Inventory', index_col='Item Desc')
    print(df_items.loc[item_name, 'On hand qty'])

    if df_items.loc[item_name, 'On hand qty'] < quantity:
        return False

    book = op.load_workbook(filename='data_files/test.xlsx')
    sheet = book['Inventory']
    sheet.cell(row=item['id'] + 2, column=13).value = df_items.loc[item_name, 'On hand qty'] - quantity
    book.save(filename='data_files/test.xlsx')
    print(df_items.loc[item_name, 'On hand qty'])
# modify("lop", 2)


# modify sales for inventory file
def setSales(id: str, quantity: int, customer: str):
    item = fetchItem(id)

    if quantity > item['available']:
        print(f"Not enough in stock, only {item['available']} are available")
        return False

    total = calculateSales(item['price'], quantity)
    sale = [{"Item Desc": id, "Price": item["price"], "Quantity": quantity, "Total": total, "Customer": customer,
             "Time": datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]
    df = pd.DataFrame(sale)

    print(total)

    rows = df.values.tolist()
    book = op.load_workbook(filename='data_files/test.xlsx')
    sheet = book['Sales']

    for row in rows:
        sheet.append(row)
    book.save(filename='data_files/test.xlsx')

    # modify inventory data
    modify_after_sale(id, quantity)

    return total
# setSales(id='glop', quantity=10, customer='lewi')


# set new items to file
def set_items(description: str, category: str, made_in: str, size: float, unit: str, quantity: int, bar_code: int,
              price: float):
    profit_margin = price * 0.45
    profit_added_price = price + (price * 0.45)
    VAT = profit_margin * 0.15
    total = VAT + profit_added_price
    item = [{"Item Desc": description, "Category": category, "Made in": made_in, "Size/ml-g-oz": size, "Unit": unit,
             "Quantity": quantity, "Bar Code": int(bar_code), "Unit Price": price, "45% Profit": "%.2f" % profit_margin,
             "Unit Price after profit": "%.2f" % profit_added_price, "VAT": "%.2f" % VAT, "Total": "%.2f" % total,
             "On hand qty": quantity, "Sold qty": "%.2f" % 0, "Total sales": "%.2f" % 0}]

    df = pd.DataFrame(item)
    rows = df.values.tolist()
    book = op.load_workbook(filename='data_files/test.xlsx')
    sheet = book['Inventory']

    for row in rows:
        sheet.append(row)
    book.save(filename='data_files/test.xlsx')
# set_items('glplpp', 'stock', 'eth', 12, 'pcs', 23, 921321546465, 10.5)


def set_i(items: dict):
    profit = float(items['Unit Price']) * 0.45
    profit_added_price = float(items['Unit Price']) + profit
    vat = profit_added_price * 0.15
    total = vat + profit_added_price

    items["45% Profit"] = "%.2f" % profit
    items["Unit Price after profit"] = "%.2f" % profit_added_price
    items["VAT"] = "%.2f" % vat
    items["Total"] = "%.2f" % total
    items["On hand qty"] = items['Quantity']
    items["Sold qty"] = 0
    items["Total sales"] = 0
    print(">>Item added", items)

    item = [items]
    df = pd.DataFrame(item)
    rows = df.values.tolist()
    book = op.load_workbook(filename='data_files/test.xlsx')
    sheet = book['Inventory']

    for row in rows:
        sheet.append(row)
    book.save(filename='data_files/test.xlsx')


# remove stock
def delete_item(item_id: str):
    item = fetchItem(item_id)
    df = pd.read_excel(io='data_files/test.xlsx', sheet_name='Inventory')
    if int(df.count()['Item Desc']) < item["id"] + 1:
        pass
    else:
        print('#' * 40)

        if item["id"] + 2 == 1 or item["id"] + 2 == 0:
            print('Invalid item')
        else:
            book = op.load_workbook(filename='data_files/test.xlsx')
            sheet = book['Inventory']
            sheet.delete_rows(idx=(item["id"] + 2), amount=1)
            book.save(filename='data_files/test.xlsx')

        print(">>", item["name"], "deleted from file\n")

    df = pd.read_excel(io='data_files/test.xlsx', sheet_name='Inventory')
    print(df)

    if df.columns.size == 0:
        initializer()
# delete_item('glop')


# edit item data
def edit(item_id: str, edited_item: dict):
    item = fetchItem(item_id)
    df_items = pd.read_excel(io='data_files/test.xlsx', sheet_name='Inventory', index_col='Item Desc')
    print(df_items.loc[item_id, 'On hand qty'])

    book = op.load_workbook(filename='data_files/test.xlsx')
    sheet = book['Inventory']
    for idx in edited_item:
        sheet.cell(row=item['id'] + 2, column=idx).value = edited_item[idx]
    book.save(filename='data_files/test.xlsx')
