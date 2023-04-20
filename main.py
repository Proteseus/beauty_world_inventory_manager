import os, signal, logging
from os.path import exists
from waitress import serve
from flask import Flask, render_template, redirect, request, flash, session, jsonify
from flask_session import Session
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
import sqlalchemy.sql.default_comparator
import sqlalchemy.dialects.sqlite
import db
from model import User
import pandas as pd
import openpyxl as op
from datetime import datetime
import webbrowser

engine = create_engine('sqlite:///instance/sample.sqlite')
app = Flask(__name__)
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_TYPE'] = "filesystem"
app.config.update(
    TESTING=True,
    SECRET_KEY="192b9bdd22ab9ed4d12e236c78afcb9a393ec15f71bbf5dc987d54727823bcbf"
)

Session(app)


@app.route('/')
def index():
    if not session.get('name'):
        return login_page()

    stock = db.fetch_all()
    return render_template('index.html', data=stock)


@app.route('/login', methods=['GET'])
def login_page():
    if not session.get('logged_in'):
        return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    form = request.form
    _session = sessionmaker(bind=engine)
    s = _session()

    user = form['username']
    password = form['password']

    if user == '' or password == '':
        flash("Fields can't be empty")
    elif db.login(user=user, password=password):
        session['logged_in'] = True
        session['role'] = db.get_role(username=user)
        session['name'] = user
        session['kart'] = []
    else:
        flash('wrong password!')
    return index()


@app.route('/logout', methods=['GET'])
def logout():
    session['logged_in'] = False
    session.pop('name')
    session.pop('role')

    for files in os.listdir('flask_session/'):
        os.remove(f'flask_session/{files}')

    db.clear_cart()

    return login_page()


@app.route('/signup', methods=['GET'])
def sign_up_page():
    return render_template('signup.html')


@app.route('/signup', methods=['POST'])
def sign_up():
    form = request.form
    if db.signup(form['username'], form['password'], 'sales'):
        return login_page()
    else:
        flash("error contact system admin")
    return index()


@app.route('/add/', methods=['POST', 'GET'])
def add():
    if not session.get('name'):
        return login_page()

    if session['role'] == 'admin':
        form = request.form
        if request.method == 'POST':
            item = []
            for idx, i_key in enumerate(form.keys()):
                item.append(form[i_key])
            print(item)
            db.set_items(item)
            return redirect('/')
    else:
        flash("Unauthorized access!")
        return index()
    return render_template('add_listing.html', form=form)


@app.route('/edit/<item_name>', methods=['GET'])
def edit(item_name: str):
    if not session.get('name'):
        return login_page()

    if session.get('role') == 'admin':
        print(item_name)
        data = db.fetch_item(item_name)
        new_data = []
        for d in data[0]:
            new_data.append(str(d))
        return render_template('edit.html', data=new_data)
    else:
        flash("Unauthorized access!")
    return index()


@app.route('/edit/', methods=['POST'])
def edit_item():
    form_data = request.form
    edited_item = []
    for idx, key in enumerate(form_data.keys()):
        edited_item.append(form_data[key])
    print(edited_item)
    db.edit_item(edited_item[1], edited_item)

    return redirect('/')


@app.route('/delete/<item_name>', methods=['GET'])
def delete(item_name: str):
    if not session.get('name'):
        return login_page()

    if session.get('role') == 'admin':
        print(item_name)
        print(type(item_name))
        db.delete_item(item_name)
        return redirect('/')
    else:
        flash("Unauthorized access!")
    return index()


#######################################
@app.route('/edit_sale/<item_id>', methods=['GET'])
def edit_sale(item_id: int):
    if not session.get('name'):
        return login_page()

    if session.get('role') == 'admin':
        print(item_id)
        data = db.fetch_sale(item_id)
        new_data = []
        for d in data[0]:
            new_data.append(str(d))
        return render_template('edit_sale.html', data=new_data, item_id=item_id)
    else:
        flash("Unauthorized access!")
    return sold()


@app.route('/edit_sale/<item_id>', methods=['POST'])
def edit_sale_item(item_id):
    form_data = request.form
    edited_item = [item_id]
    for idx, key in enumerate(form_data.keys()):
        edited_item.append(form_data[key])
    print(edited_item)
    db.edit_sale(edited_item[0], edited_item)

    return sold()


@app.route('/delete_sale/<item_name>', methods=['GET'])
def delete_sale(item_name: str):
    if not session.get('name'):
        return login_page()

    if session.get('role') == 'admin':
        print(item_name)
        print(type(item_name))
        param = [item_name.split(':')[0], item_name.split(':')[1][0:]]
        print(param)
        print(type(param))
        db.delete_sale(param)
        return sold()
    else:
        flash("Unauthorized access!")
    return sold()


##########################


@app.route("/get_data")
def get_data():
    input_value = request.args.get("input")
    print(input_value)
    # fetch data based on input_value
    conn, cur = db.connector()
    cur.execute('SELECT item FROM inventory WHERE barcode=?', (input_value,))
    data = cur.fetchone()
    cur.close()
    conn.close()

    return data[0]


@app.route('/sales/', methods=['GET'])
def sale():
    if not session.get('name'):
        return login_page()

    item_names = []
    barcodes = []
    items = db.fetch_all()

    for item in items:
        # print(item)
        item_names.append(str(item[0]))
        barcodes.append(item[6])

    kart_items = db.fetch_cart()
    total = 0
    for items in kart_items:
        total += items[2]

    return render_template('sales.html', items=item_names, barcodes=barcodes, kart=kart_items, total=total)


@app.route('/sales/', methods=['POST'])
def sale_item():
    form_data = request.form
    if form_data.get('name') == '' or form_data.get('quantity') == '':
        return render_template('sales.html')
    result, remaining = db.on_hand_checker(form_data['name'], int(form_data['quantity']))
    if result:
        kart_item = [form_data['name'], int(form_data['quantity']), form_data['customer'], form_data['phone_number']]
        add_to_cart(kart_item)
    else:
        flash(f"Not enough in stock, only {remaining} are available", 'warning')
    return redirect('/sales/')


def add_to_cart(items):
    db.add_to_kart(items)
    return sale()


@app.route('/remove_from_cart/<item_name>', methods=['GET'])
def remove_from_cart(item_name: str):
    db.remove_from_cart(item_name)
    return sale()


@app.route('/checkout', methods=['GET'])
def checkout():
    form_data = request.form
    kart_items = db.fetch_cart()
    print(request.args.get('name'))
    for items in kart_items:
        db.set_sale(items[0], int(items[1]), form_data.get('name'), form_data.get('phone'))
    return sold()


@app.route('/sold/', methods=['GET'])
def sold():
    if not session.get('name'):
        return login_page()

    data = db.fetch_all_sold()
    return render_template('all_sold.html', data=data)


@app.route('/supplier', methods=['GET'])
def supplier_page():
    if not session.get('name'):
        return login_page()

    suppliers = db.fetch_supplier()
    return render_template('supplier.html', suppliers=suppliers)


@app.route('/add_supplier', methods=['GET'])
def supplier_app():
    if not session.get('name'):
        return login_page()
    return render_template('add_supplier.html')


@app.route('/add_supplier', methods=['POST'])
def supplier():
    if not session.get('name'):
        return login_page()

    form_data = request.form
    suppliers = []
    for idx, key in enumerate(form_data.keys()):
        suppliers.append(form_data[key])
    db.set_supplier(suppliers)
    return supplier_page()


@app.route('/report', methods=['GET'])
def report():
    if not session.get('name'):
        return login_page()

    return render_template('reports.html')


@app.route('/daily_report', methods=['GET'])
def daily_report():
    rows = db.get_daily_report()
    df = pd.DataFrame(rows,
                      columns=['id', 'item', 'category', 'made in', 'size/ml-g-oz', 'unit', 'quantity', 'barcode',
                               'price', 'profit', 'price after profit', 'total', 'vat', 'customer', 'phone_number',
                               'date'])
    df.set_index('id', inplace=True)
    print(df['item'].to_string())

    rows = df.values.tolist()
    books = op.Workbook()
    if not exists('reports/daily.xlsx'):
        books.save('reports/daily.xlsx')
    book = op.load_workbook(filename='reports/daily.xlsx')
    sheet = book.active
    if sheet.title != datetime.now().strftime('%Y-%m-%d'):
        book.create_sheet(title=datetime.now().strftime('%Y-%m-%d'))
        cols = [{'item': '', 'category': '', 'made in': '', 'size/ml-g-oz': '', 'unit': '', 'quantity': '',
                 'barcode': '', 'price': '', 'profit': '', 'price after profit': '', 'total': '', 'vat': '',
                 'customer': '', 'phone_number': '', 'date': ''}]

        for ite in cols[0].keys():
            cols[0][ite] = ite

        cols = pd.DataFrame(cols).values.tolist()
        sheet = book[datetime.now().strftime('%Y-%m-%d')]
        for row in cols:
            sheet.append(row)
        # book.remove_sheet('Sheet1')
        book.save('reports/daily.xlsx')
    else:
        sheet = book[datetime.now().strftime('%Y-%m-%d')]
    for row in rows:
        sheet.append(row)
    book.save(filename='reports/daily.xlsx')
    return report()


@app.route('/weekly_report', methods=['GET'])
def weekly_report():
    rows = db.get_weekly_report()
    df = pd.DataFrame(rows,
                      columns=['id', 'item', 'category', 'made in', 'size/ml-g-oz', 'unit', 'quantity', 'barcode',
                               'price', 'profit', 'price after profit', 'total', 'vat', 'customer', 'phone_number',
                               'date'])
    df.set_index('id', inplace=True)
    print(df['item'].to_string())

    rows = df.values.tolist()
    books = op.Workbook()
    if not exists('reports/weekly.xlsx'):
        books.save('reports/weekly.xlsx')
    book = op.load_workbook(filename='reports/weekly.xlsx')
    sheet = book.active
    if sheet.title != str(datetime.now().isocalendar().week):
        book.create_sheet(title=str(datetime.now().isocalendar().week))
        cols = [{'item': '', 'category': '', 'made in': '', 'size/ml-g-oz': '', 'unit': '', 'quantity': '',
                 'barcode': '', 'price': '', 'profit': '', 'price after profit': '', 'total': '', 'vat': '',
                 'customer': '', 'phone_number': '', 'date': ''}]

        for ite in cols[0].keys():
            cols[0][ite] = ite

        cols = pd.DataFrame(cols).values.tolist()
        sheet = book[str(datetime.now().isocalendar().week)]
        for row in cols:
            sheet.append(row)
        # book.remove_sheet('Sheet1')
        book.save('reports/weekly.xlsx')
    else:
        sheet = book[str(datetime.now().isocalendar().week)]
    for row in rows:
        sheet.append(row)
    book.save(filename='reports/weekly.xlsx')
    return report()


@app.route('/monthly_report', methods=['GET'])
def monthly_report():
    rows = db.get_monthly_report()
    df = pd.DataFrame(rows,
                      columns=['id', 'item', 'category', 'made in', 'size/ml-g-oz', 'unit', 'quantity', 'barcode',
                               'price', 'profit', 'price after profit', 'total', 'vat', 'customer', 'phone_number',
                               'date'])
    df.set_index('id', inplace=True)
    print(df['item'].to_string())

    rows = df.values.tolist()
    books = op.Workbook()
    if not exists('reports/monthly.xlsx'):
        books.save('reports/monthly.xlsx')
    book = op.load_workbook(filename='reports/monthly.xlsx')
    sheet = book.active
    if sheet.title != datetime.now().strftime('%Y-%m'):
        book.create_sheet(title=datetime.now().strftime('%Y-%m'))
        cols = [{'item': '', 'category': '', 'made in': '', 'size/ml-g-oz': '', 'unit': '', 'quantity': '',
                 'barcode': '', 'price': '', 'profit': '', 'price after profit': '', 'total': '', 'vat': '',
                 'customer': '', 'phone_number': '', 'date': ''}]

        for ite in cols[0].keys():
            cols[0][ite] = ite

        cols = pd.DataFrame(cols).values.tolist()
        sheet = book[datetime.now().strftime('%Y-%m')]
        for row in cols:
            sheet.append(row)
        # book.remove_sheet('Sheet1')
        book.save('reports/monthly.xlsx')
    else:
        sheet = book[datetime.now().strftime('%Y-%m')]
    for row in rows:
        sheet.append(row)
    book.save(filename='reports/monthly.xlsx')
    return report()


@app.route('/annual_report', methods=['GET'])
def annual_report():
    rows = db.get_annual_report()
    df = pd.DataFrame(rows,
                      columns=['id', 'item', 'category', 'made in', 'size/ml-g-oz', 'unit', 'quantity', 'barcode',
                               'price', 'profit', 'price after profit', 'total', 'vat', 'customer', 'phone_number',
                               'date'])
    df.set_index('id', inplace=True)
    print(df['item'].to_string())

    rows = df.values.tolist()
    books = op.Workbook()
    if not exists('reports/annual.xlsx'):
        books.save('reports/annual.xlsx')
    book = op.load_workbook(filename='reports/annual.xlsx')
    sheet = book.active
    if sheet.title != datetime.now().strftime('%Y'):
        book.create_sheet(title=datetime.now().strftime('%Y'))
        cols = [{'item': '', 'category': '', 'made in': '', 'size/ml-g-oz': '', 'unit': '', 'quantity': '',
                 'barcode': '', 'price': '', 'profit': '', 'price after profit': '', 'total': '', 'vat': '',
                 'customer': '', 'phone_number': '', 'date': ''}]

        for ite in cols[0].keys():
            cols[0][ite] = ite

        cols = pd.DataFrame(cols).values.tolist()
        sheet = book[datetime.now().strftime('%Y')]
        for row in cols:
            sheet.append(row)
        # book.remove_sheet('Sheet1')
        book.save('reports/annual.xlsx')
    else:
        sheet = book[datetime.now().strftime('%Y')]
    for row in rows:
        sheet.append(row)
    book.save(filename='reports/annual.xlsx')
    return report()


@app.route('/search', methods=['POST'])
def search():
    form_data = request.form
    search_result = db.fetch_item_sim(form_data.get('search-item'))
    return render_template('index.html', data=search_result)


def expiry_date_notify():
    pass


@app.route('/stopServer', methods=['GET'])
def stop_server():
    if 'logged_in' in session.keys():
        logout()
    os.kill(os.getpid(), signal.SIGINT)
    return jsonify({"success": True, "message": "Server is shutting down..."})


if __name__ == '__main__':
    # app.run(debug=True)
    webbrowser.open('http://localhost:8081/', new=2)
    serve(app, host='0.0.0.0', port=8081)

